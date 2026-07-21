"""
price_tracker.py
----------------
Tracks product prices across Amazon, Flipkart, Snapdeal, Meesho, Blinkit.
Saves history to D:\\price_tracker.xlsx and alerts when price drops.

Usage:
    python price_tracker.py                  # track all items in items.json
    python price_tracker.py --add            # add a new item interactively
    python price_tracker.py --list           # show all tracked items + latest price
    python price_tracker.py --interval 60   # check every 60 minutes (default: 120)
"""

import os
import re
import sys
import json
import time
import random
import argparse
import threading
from datetime import datetime
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook

DEFAULT_XLS    = r"D:\price_tracker.xlsx"
ITEMS_JSON     = os.path.join(os.path.dirname(__file__), "items.json")
CHECK_INTERVAL = 120   # minutes between checks

# Rotate user agents to avoid blocks
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4.1 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 Edg/123.0.0.0",
]

def get_headers():
    return {
        "User-Agent": random.choice(USER_AGENTS),
        "Accept-Language": "en-IN,en;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "DNT": "1",
    }

# ─── Items Store ─────────────────────────────────────────────────────────────

def load_items():
    if not os.path.exists(ITEMS_JSON):
        # Default sample items
        items = [
            {
                "name":       "Devil May Cry Pen",
                "keywords":   "devil may cry pen",
                "sites":      ["amazon", "flipkart", "snapdeal", "meesho"],
                "target_price": 500,
                "alert":      True
            }
        ]
        save_items(items)
        return items
    with open(ITEMS_JSON) as f:
        return json.load(f)

def save_items(items):
    with open(ITEMS_JSON, "w") as f:
        json.dump(items, f, indent=2)

# ─── Scrapers ─────────────────────────────────────────────────────────────────

def clean_price(text):
    """Extract numeric price from strings like '₹1,299', 'Rs. 499', etc."""
    if not text:
        return None
    text = re.sub(r"[^\d.]", "", text.replace(",", ""))
    try:
        p = float(text) if text else None
        return p if p and p > 1 else None   # filter out Rs.0 / junk
    except Exception:
        return None


def scrape_amazon(keyword):
    """Search Amazon India and return list of {name, price, url, site}."""
    results = []
    try:
        query = keyword.replace(" ", "+")
        url   = f"https://www.amazon.in/s?k={query}"
        resp  = requests.get(url, headers=get_headers(), timeout=15)
        soup  = BeautifulSoup(resp.text, "html.parser")

        for item in soup.select('[data-component-type="s-search-result"]')[:5]:
            try:
                name_el  = item.select_one("h2 span.a-text-normal, h2 a span")
                price_el = item.select_one(".a-price .a-offscreen")
                link_el  = item.select_one("h2 a")

                name  = name_el.text.strip()  if name_el  else "Unknown"
                price = clean_price(price_el.text) if price_el else None
                link  = "https://www.amazon.in" + link_el["href"] if link_el else url

                if price:
                    results.append({"name": name[:80], "price": price,
                                    "url": link, "site": "Amazon"})
            except Exception:
                continue
    except Exception as e:
        print(f"  [Amazon] Error: {e}")
    return results


def scrape_flipkart(keyword):
    results = []
    try:
        query = keyword.replace(" ", "%20")
        url   = f"https://www.flipkart.com/search?q={query}"
        resp  = requests.get(url, headers=get_headers(), timeout=15)
        soup  = BeautifulSoup(resp.text, "html.parser")

        for item in soup.select("._1AtVbE")[:8]:
            try:
                name_el  = item.select_one("._4rR01T, .s1Q9rs, .IRpwTa")
                price_el = item.select_one("._30jeq3, ._1_WHN1")
                link_el  = item.select_one("a._1fQZEK, a.s1Q9rs, a._2rpwqI")

                name  = name_el.text.strip()  if name_el  else None
                price = clean_price(price_el.text) if price_el else None
                link  = "https://www.flipkart.com" + link_el["href"] if link_el else url

                if name and price:
                    results.append({"name": name[:80], "price": price,
                                    "url": link, "site": "Flipkart"})
            except Exception:
                continue
    except Exception as e:
        print(f"  [Flipkart] Error: {e}")
    return results


def scrape_snapdeal(keyword):
    results = []
    try:
        query = keyword.replace(" ", "%20")
        url   = f"https://www.snapdeal.com/search?keyword={query}"
        resp  = requests.get(url, headers=get_headers(), timeout=15)
        soup  = BeautifulSoup(resp.text, "html.parser")

        for item in soup.select(".product-tuple-listing")[:5]:
            try:
                name_el  = item.select_one(".product-title")
                price_el = item.select_one(".product-price")
                link_el  = item.select_one("a.dp-widget-link")

                name  = name_el.text.strip()  if name_el  else None
                price = clean_price(price_el.text) if price_el else None
                link  = link_el["href"]             if link_el  else url

                if name and price:
                    results.append({"name": name[:80], "price": price,
                                    "url": link, "site": "Snapdeal"})
            except Exception:
                continue
    except Exception as e:
        print(f"  [Snapdeal] Error: {e}")
    return results


def scrape_meesho(keyword):
    results = []
    try:
        query = keyword.replace(" ", "%20")
        url   = f"https://www.meesho.com/search?q={query}"
        resp  = requests.get(url, headers=get_headers(), timeout=15)
        soup  = BeautifulSoup(resp.text, "html.parser")

        for item in soup.select("[class*='ProductList__GridCol']")[:5]:
            try:
                name_el  = item.select_one("p[class*='Text']")
                price_el = item.select_one("h5[class*='Text']")
                link_el  = item.select_one("a")

                name  = name_el.text.strip()  if name_el  else None
                price = clean_price(price_el.text) if price_el else None
                link  = "https://www.meesho.com" + link_el["href"] if link_el else url

                if name and price:
                    results.append({"name": name[:80], "price": price,
                                    "url": link, "site": "Meesho"})
            except Exception:
                continue
    except Exception as e:
        print(f"  [Meesho] Error: {e}")
    return results


def scrape_blinkit(keyword):
    """Blinkit requires JS — returns search URL for manual check."""
    results = []
    try:
        query = keyword.replace(" ", "%20")
        url   = f"https://blinkit.com/s/?q={query}"
        # Blinkit is React-rendered; note the URL for reference
        results.append({
            "name":  f"Search '{keyword}' on Blinkit",
            "price": None,
            "url":   url,
            "site":  "Blinkit",
            "note":  "JS-rendered — open URL manually"
        })
    except Exception as e:
        print(f"  [Blinkit] Error: {e}")
    return results


SCRAPERS = {
    "amazon":   scrape_amazon,
    "flipkart": scrape_flipkart,
    "snapdeal": scrape_snapdeal,
    "meesho":   scrape_meesho,
    "blinkit":  scrape_blinkit,
}


def search_item(item):
    """Run all configured scrapers for an item, return combined results."""
    keyword  = item["keywords"]
    sites    = item.get("sites", list(SCRAPERS.keys()))
    all_results = []

    for site in sites:
        if site in SCRAPERS:
            print(f"    Searching {site.capitalize()}...")
            results = SCRAPERS[site](keyword)
            all_results.extend(results)
            time.sleep(random.uniform(1.5, 3.0))  # polite delay

    return all_results


# ─── Excel ────────────────────────────────────────────────────────────────────

_lock = threading.Lock()

def init_excel(filepath):
    if not os.path.exists(filepath):
        wb = openpyxl.Workbook()

        # Sheet 1: Price History
        ws = wb.active
        ws.title = "Price History"
        headers  = ["#", "Checked At", "Item", "Product Name", "Site",
                    "Price (INR)", "Target Price", "Status", "URL"]
        ws.append(headers)
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill; cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 25
        ws.column_dimensions["D"].width = 45
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 14
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 60

        # Sheet 2: Best Prices (lowest per item per site)
        ws2 = wb.create_sheet("Best Prices")
        ws2.append(["Item", "Best Price (INR)", "Site", "Product", "URL", "Last Updated"])
        fill2 = PatternFill("solid", fgColor="375623")
        font2 = Font(color="FFFFFF", bold=True)
        for cell in ws2[1]:
            cell.fill = fill2; cell.font = font2
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 12
        ws2.column_dimensions["D"].width = 45
        ws2.column_dimensions["E"].width = 60
        ws2.column_dimensions["F"].width = 20

        wb.save(filepath)
        print(f"Created: {filepath}")
    else:
        wb   = load_workbook(filepath)
        rows = wb["Price History"].max_row - 1
        print(f"Appending to: {filepath} ({rows} existing records)")
    return load_workbook(filepath)


def save_results(wb, filepath, item, results):
    ts     = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    target = item.get("target_price")

    with _lock:
        ws  = wb["Price History"]
        ws2 = wb["Best Prices"]

        for r in results:
            price  = r.get("price")
            if price is None:
                continue

            row    = ws.max_row + 1
            status = ""
            color  = "FFFFFF"

            if target and price <= target:
                status = "BELOW TARGET"
                color  = "C6EFCE"  # green
            elif target and price <= target * 1.1:
                status = "NEAR TARGET"
                color  = "FFEB9C"  # yellow
            else:
                status = "Above Target" if target else "Tracked"
                color  = "FFFFFF"

            ws.append([
                row - 1, ts,
                item["name"], r["name"], r["site"],
                price, target or "", status, r["url"]
            ])

            fill = PatternFill("solid", fgColor=color)
            for col in range(1, 10):
                ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=9).alignment = Alignment(wrap_text=True)

        # Update Best Prices sheet
        _update_best_prices(ws2, item, results, ts)
        wb.save(filepath)


def _update_best_prices(ws2, item, results, ts):
    priced = [r for r in results if r.get("price")]
    if not priced:
        return
    best = min(priced, key=lambda x: x["price"])

    # Update or insert
    for row in ws2.iter_rows(min_row=2):
        if row[0].value == item["name"]:
            if best["price"] < (row[1].value or 99999):
                row[1].value = best["price"]
                row[2].value = best["site"]
                row[3].value = best["name"]
                row[4].value = best["url"]
                row[5].value = ts
                # Highlight best price in gold
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="FFD700")
            return

    ws2.append([item["name"], best["price"], best["site"],
                best["name"], best["url"], ts])


# ─── Notifications ────────────────────────────────────────────────────────────

def notify_price_drop(item, result):
    try:
        from winotify import Notification
        Notification(
            app_id = "Price Tracker",
            title  = f"Price Drop! {item['name']}",
            msg    = f"Rs.{result['price']} on {result['site']}\n(Target: Rs.{item.get('target_price','?')})\n{result['name'][:60]}",
            duration = "long"
        ).show()
    except Exception:
        pass


# ─── Main Tracker Loop ────────────────────────────────────────────────────────

def run_once(wb, filepath):
    items = load_items()
    ts    = datetime.now().strftime("%H:%M:%S")
    print(f"\n[{ts}] Checking {len(items)} item(s)...\n")

    for item in items:
        print(f"  [{item['name']}]")
        results = search_item(item)

        if results:
            save_results(wb, filepath, item, results)
            priced = [r for r in results if r.get("price")]
            for r in priced:
                status = ""
                if item.get("target_price") and r["price"] <= item["target_price"]:
                    status = " <-- BELOW TARGET!"
                    if item.get("alert"):
                        notify_price_drop(item, r)
                print(f"    {r['site']:10} Rs.{r['price']:>8,.0f}  {r['name'][:45]}{status}")
        else:
            print(f"    No results found.")
        print()


def cmd_list():
    items = load_items()
    if not os.path.exists(DEFAULT_XLS):
        print("No price data yet. Run the tracker first.")
        return
    wb = load_workbook(DEFAULT_XLS)
    ws = wb["Best Prices"]
    print(f"\n{'='*70}")
    print(f"  TRACKED ITEMS — BEST PRICES FOUND")
    print(f"{'='*70}")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            price = f"Rs.{row[1]:,.0f}" if row[1] else "No data"
            print(f"  {str(row[0])[:30]:<30} {price:>12}  ({row[2] or '-'})")
    print()


def cmd_add():
    items   = load_items()
    name    = input("Item name (e.g. Devil May Cry Pen): ").strip()
    kw      = input("Search keywords: ").strip() or name
    sites   = input("Sites [amazon,flipkart,snapdeal,meesho,blinkit]: ").strip()
    sites   = [s.strip() for s in sites.split(",")] if sites else ["amazon","flipkart","snapdeal","meesho"]
    target  = input("Target price (INR, press Enter to skip): ").strip()
    target  = float(target) if target else None

    items.append({
        "name":         name,
        "keywords":     kw,
        "sites":        sites,
        "target_price": target,
        "alert":        True
    })
    save_items(items)
    print(f"\nAdded '{name}'. Run the tracker to fetch prices.\n")


# ─── Entry Point ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Price Tracker")
    parser.add_argument("--list",     action="store_true", help="Show best prices found")
    parser.add_argument("--add",      action="store_true", help="Add a new item to track")
    parser.add_argument("--once",     action="store_true", help="Check prices once and exit")
    parser.add_argument("--interval", default=CHECK_INTERVAL, type=int,
                        help=f"Check interval in minutes (default: {CHECK_INTERVAL})")
    parser.add_argument("--output",   default=DEFAULT_XLS, help="Excel output path")
    args = parser.parse_args()

    if args.list:
        cmd_list(); return
    if args.add:
        cmd_add();  return

    print("\n Price Tracker")
    print("=" * 50)
    wb = init_excel(args.output)

    if args.once:
        run_once(wb, args.output)
        return

    print(f"Checking every {args.interval} minutes. Press Ctrl+C to stop.\n")
    while True:
        run_once(wb, args.output)
        print(f"Next check in {args.interval} minutes...\n")
        time.sleep(args.interval * 60)


if __name__ == "__main__":
    main()
