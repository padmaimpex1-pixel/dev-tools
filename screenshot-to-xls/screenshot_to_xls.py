"""
screenshot_to_xls.py
--------------------
Monitors clipboard (PrtScn) and Screenshots folder (Win+PrtScn).
Extracts text via OCR and saves everything to an Excel file.
Also captures open Notepad windows and Windows Sticky Notes on each screenshot.

Usage:
    python screenshot_to_xls.py
    python screenshot_to_xls.py --output my_data.xlsx
    python screenshot_to_xls.py --folder "C:/Users/YourName/Pictures/Screenshots"
"""

import os
import sys
import time
import sqlite3
import hashlib
import argparse
import threading
from datetime import datetime
import re

from PIL import ImageGrab, Image
import pytesseract
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# ─── Config ──────────────────────────────────────────────────────────────────

DEFAULT_XLS      = r"D:\screenshots_data.xlsx"
DEFAULT_SHOTS_DIR = os.path.expanduser("~/Pictures/Screenshots")
STICKY_NOTES_DB = os.path.expandvars(
    r"%LOCALAPPDATA%\Packages\Microsoft.MicrosoftStickyNotes_8wekyb3d8bbwe\LocalState\plum.sqlite"
)

# ─── Excel Setup ─────────────────────────────────────────────────────────────

def init_excel(filepath):
    """Create Excel file with styled headers if it doesn't exist."""
    if not os.path.exists(filepath):
        wb = openpyxl.Workbook()

        # ── Sheet 1: Screenshots ──
        ws = wb.active
        ws.title = "Screenshots"
        headers = ["#", "Timestamp", "Source", "Filename", "Extracted Text", "Open Notepads", "Sticky Notes"]
        ws.append(headers)
        _style_header(ws, len(headers))
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 30
        ws.column_dimensions["E"].width = 60
        ws.column_dimensions["F"].width = 40
        ws.column_dimensions["G"].width = 40
        ws.row_dimensions[1].height = 20

        # ── Sheet 2: Sticky Notes ──
        ws2 = wb.create_sheet("Sticky Notes")
        ws2.append(["#", "Captured At", "Note ID", "Content"])
        _style_header(ws2, 4)
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 40
        ws2.column_dimensions["D"].width = 80

        # ── Sheet 3: Notepad ──
        ws3 = wb.create_sheet("Notepad")
        ws3.append(["#", "Captured At", "Window Title", "Content"])
        _style_header(ws3, 4)
        ws3.column_dimensions["A"].width = 6
        ws3.column_dimensions["B"].width = 22
        ws3.column_dimensions["C"].width = 40
        ws3.column_dimensions["D"].width = 80

        wb.save(filepath)
        print(f"Created Excel file: {filepath}")
    else:
        _wb = load_workbook(filepath)
        existing_rows = _wb["Screenshots"].max_row - 1
        print(f"Appending to existing file: {filepath} ({existing_rows} existing entries)")

    return load_workbook(filepath)


def _style_header(ws, num_cols):
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1][:num_cols]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


# ─── Excel Write (thread-safe) ────────────────────────────────────────────────

_excel_lock = threading.Lock()

def save_to_excel(wb, filepath, source, filename, text):
    """Append a new row with OCR text + live Notepad + Sticky Notes snapshot."""
    notepads = get_notepad_contents()
    stickies = get_sticky_notes()
    ts       = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    notepad_summary = " | ".join(
        f"[{t}]: {c[:60]}..." if len(c) > 60 else f"[{t}]: {c}"
        for t, c in notepads
    ) or "—"

    sticky_summary = " | ".join(
        s[:80] + "..." if len(s) > 80 else s
        for s in stickies
    ) or "—"

    with _excel_lock:
        # ── Sheet 1: Screenshots ──
        ws  = wb["Screenshots"]
        row = ws.max_row + 1
        ws.append([row - 1, ts, source, filename, text, notepad_summary, sticky_summary])
        for col in [5, 6, 7]:
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True)
        if row % 2 == 0:
            fill = PatternFill("solid", fgColor="DEEAF1")
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = fill

        # ── Sheet 2: Sticky Notes (each note as a row) ──
        ws2 = wb["Sticky Notes"]
        for note_id, content in enumerate(stickies, ws2.max_row):
            ws2.append([note_id, ts, f"note_{note_id}", content])
            ws2.cell(row=ws2.max_row, column=4).alignment = Alignment(wrap_text=True)

        # ── Sheet 3: Notepad (each window as a row) ──
        ws3 = wb["Notepad"]
        for title, content in notepads:
            r = ws3.max_row + 1
            ws3.append([r - 1, ts, title, content])
            ws3.cell(row=r, column=4).alignment = Alignment(wrap_text=True)

        wb.save(filepath)
        print(f"[{ts}] Saved -> {source} | {filename} | OCR:{len(text)}ch | "
              f"Notepads:{len(notepads)} | Stickies:{len(stickies)}")



# ─── Notepad Reader ───────────────────────────────────────────────────────────

def get_notepad_contents():
    """Return list of (window_title, text_content) for all open Notepad windows."""
    results = []
    try:
        import win32gui
        import win32con

        def enum_callback(hwnd, _):
            if not win32gui.IsWindowVisible(hwnd):
                return
            title = win32gui.GetWindowText(hwnd)
            # Match Notepad, Notepad++, and plain .txt windows
            if "Notepad" in title or title.endswith(".txt"):
                # Get the Edit child control
                edit_hwnd = win32gui.FindWindowEx(hwnd, 0, "Edit", None)
                if not edit_hwnd:
                    edit_hwnd = win32gui.FindWindowEx(hwnd, 0, "RichEditD2DPT", None)
                if edit_hwnd:
                    length   = win32gui.SendMessage(edit_hwnd, win32con.WM_GETTEXTLENGTH, 0, 0)
                    if length > 0:
                        import ctypes
                        buf = ctypes.create_unicode_buffer(length + 1)
                        ctypes.windll.user32.SendMessageW(edit_hwnd, win32con.WM_GETTEXT, length + 1, buf)
                        results.append((title, buf.value.strip()))

        win32gui.EnumWindows(enum_callback, None)
    except ImportError:
        pass
    except Exception as e:
        print(f"Notepad read error: {e}")
    return results


# ─── Sticky Notes Reader ──────────────────────────────────────────────────────

def get_sticky_notes():
    """Read all notes from Windows Sticky Notes SQLite database."""
    notes = []
    if not os.path.exists(STICKY_NOTES_DB):
        return notes
    try:
        # Copy DB to temp (it may be locked)
        import shutil, tempfile
        tmp = os.path.join(tempfile.gettempdir(), "plum_copy.sqlite")
        shutil.copy2(STICKY_NOTES_DB, tmp)

        conn = sqlite3.connect(tmp)
        cur  = conn.cursor()
        # Try both schema versions
        try:
            cur.execute("SELECT Text FROM Note WHERE IsDeleted = 0")
        except sqlite3.OperationalError:
            cur.execute("SELECT Text FROM Note")
        rows = cur.fetchall()
        conn.close()
        os.remove(tmp)

        for (raw,) in rows:
            if raw:
                # Strip RTF-like markup tags (\\id0... \\bold etc.)
                clean = re.sub(r'\\\w+\d*', ' ', raw)
                clean = re.sub(r'\s+', ' ', clean).strip()
                if clean:
                    notes.append(clean)
    except Exception as e:
        print(f"Sticky Notes read error: {e}")
    return notes


# ─── OCR ─────────────────────────────────────────────────────────────────────

def run_ocr(img):
    """Extract text from a PIL Image."""
    try:
        text = pytesseract.image_to_string(img, lang="eng").strip()
        return text if text else "[No text detected]"
    except Exception as e:
        return f"[OCR error: {e}]"


# ─── Clipboard Monitor ────────────────────────────────────────────────────────

def monitor_clipboard(wb, filepath):
    """Watch clipboard for new screenshot images (PrtScn key)."""
    last_hash = None
    print("📋 Watching clipboard (PrtScn)...")

    while True:
        try:
            img = ImageGrab.grabclipboard()
            if isinstance(img, Image.Image):
                img_hash = hashlib.md5(img.tobytes()).hexdigest()
                if img_hash != last_hash:
                    last_hash = img_hash
                    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"clipboard_{ts}.png"
                    text     = run_ocr(img)
                    save_to_excel(wb, filepath, "Clipboard (PrtScn)", filename, text)
        except Exception:
            pass
        time.sleep(1)


# ─── Folder Watcher ───────────────────────────────────────────────────────────

class ScreenshotHandler(FileSystemEventHandler):
    def __init__(self, wb, filepath):
        self.wb       = wb
        self.filepath = filepath

    def on_created(self, event):
        if event.is_directory:
            return
        path = event.src_path
        if not path.lower().endswith((".png", ".jpg", ".jpeg", ".bmp")):
            return
        time.sleep(0.8)  # wait for OS to finish writing the file
        try:
            img      = Image.open(path)
            text     = run_ocr(img)
            filename = os.path.basename(path)
            save_to_excel(self.wb, self.filepath, "Screenshots Folder", filename, text)
        except Exception as e:
            print(f"⚠️  Could not process {path}: {e}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Screenshot → Excel (OCR)")
    parser.add_argument("--output", default=DEFAULT_XLS,        help="Output Excel file path")
    parser.add_argument("--folder", default=DEFAULT_SHOTS_DIR,  help="Screenshots folder to watch")
    args = parser.parse_args()

    print("\n🖥️  Screenshot → Excel Monitor")
    print("=" * 45)

    # Check Tesseract
    try:
        pytesseract.get_tesseract_version()
        print("✅ Tesseract OCR detected")
    except Exception:
        print("❌ Tesseract not found!")
        print("   Download from: https://github.com/UB-Mannheim/tesseract/wiki")
        print(f"   Install to: {TESSERACT_PATH}")
        sys.exit(1)

    wb       = init_excel(args.output)
    observer = None

    # Start clipboard monitor thread
    clip_thread = threading.Thread(
        target=monitor_clipboard,
        args=(wb, args.output),
        daemon=True
    )
    clip_thread.start()

    # Start folder watcher
    if os.path.exists(args.folder):
        handler  = ScreenshotHandler(wb, args.output)
        observer = Observer()
        observer.schedule(handler, args.folder, recursive=False)
        observer.start()
        print(f"📁 Watching folder: {args.folder}")
    else:
        print(f"⚠️  Screenshots folder not found: {args.folder}")
        print("    Only clipboard monitoring is active.")

    print(f"📊 Saving to: {os.path.abspath(args.output)}")
    print("\nPress PrtScn or Win+PrtScn to capture. Press Ctrl+C to stop.\n")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        print("\n🛑 Stopped.")
        if observer:
            observer.stop()
            observer.join()


if __name__ == "__main__":
    main()
