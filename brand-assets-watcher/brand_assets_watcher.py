"""
brand_assets_watcher.py
-----------------------
Scans ALL drives for folders named "brand assets" (case-insensitive),
watches them for new additions, and logs every change to Excel.

Usage:
    python brand_assets_watcher.py
    python brand_assets_watcher.py --output D:\brand_assets_log.xlsx
    python brand_assets_watcher.py --rescan 60   (re-scan drives every 60 mins)
"""

import os
import sys
import time
import string
import argparse
import threading
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

DEFAULT_XLS    = r"D:\brand_assets_log.xlsx"
FOLDER_NAME    = "brand assets"   # case-insensitive match
RESCAN_MINUTES = 30               # re-scan drives for new "brand assets" folders

# File type → color mapping for Excel rows
FILE_COLORS = {
    # Images
    ".jpg":  "FFF2CC", ".jpeg": "FFF2CC", ".png":  "FFF2CC",
    ".gif":  "FFF2CC", ".webp": "FFF2CC", ".svg":  "FFF2CC",
    ".bmp":  "FFF2CC", ".tiff": "FFF2CC", ".ico":  "FFF2CC",
    # Videos
    ".mp4":  "DEEBF7", ".mov":  "DEEBF7", ".avi":  "DEEBF7",
    ".mkv":  "DEEBF7", ".wmv":  "DEEBF7",
    # Design files
    ".psd":  "E2EFDA", ".ai":   "E2EFDA", ".xd":   "E2EFDA",
    ".fig":  "E2EFDA", ".sketch":"E2EFDA", ".cdr":  "E2EFDA",
    ".eps":  "E2EFDA", ".indd": "E2EFDA",
    # Fonts
    ".ttf":  "FCE4D6", ".otf":  "FCE4D6", ".woff": "FCE4D6",
    ".woff2":"FCE4D6",
    # Documents
    ".pdf":  "F2F2F2", ".docx": "F2F2F2", ".pptx": "F2F2F2",
}

# ─── Excel ────────────────────────────────────────────────────────────────────

_lock = threading.Lock()

def init_excel(filepath):
    if not os.path.exists(filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Brand Assets"
        headers  = ["#", "Timestamp", "Event", "File Name", "Extension",
                    "Size (KB)", "Drive", "Full Path", "Watched Folder"]
        ws.append(headers)

        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill      = fill
            cell.font      = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 35
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 8
        ws.column_dimensions["H"].width = 70
        ws.column_dimensions["I"].width = 50
        ws.row_dimensions[1].height = 20

        # ── Sheet 2: Watched Folders ──
        ws2 = wb.create_sheet("Watched Folders")
        ws2.append(["#", "Drive", "Folder Path", "First Seen", "File Count"])
        fill2 = PatternFill("solid", fgColor="375623")
        font2 = Font(color="FFFFFF", bold=True)
        for cell in ws2[1]:
            cell.fill = fill2; cell.font = font2
            cell.alignment = Alignment(horizontal="center")
        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 8
        ws2.column_dimensions["C"].width = 70
        ws2.column_dimensions["D"].width = 22
        ws2.column_dimensions["E"].width = 12

        wb.save(filepath)
        print(f"Created: {filepath}")
    else:
        wb   = load_workbook(filepath)
        rows = wb["Brand Assets"].max_row - 1
        print(f"Appending to: {filepath} ({rows} existing entries)")
    return load_workbook(filepath)


def save_event(wb, filepath, event_type, file_path, watched_folder):
    p    = Path(file_path)
    ext  = p.suffix.lower()
    name = p.name
    try:
        size_kb = round(p.stat().st_size / 1024, 2) if p.exists() else 0
    except Exception:
        size_kb = 0
    drive = os.path.splitdrive(file_path)[0] or "?"
    ts    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with _lock:
        ws  = wb["Brand Assets"]
        row = ws.max_row + 1
        ws.append([row - 1, ts, event_type, name, ext,
                   size_kb, drive, file_path, watched_folder])

        color = FILE_COLORS.get(ext, "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=8).alignment = Alignment(wrap_text=True)

        wb.save(filepath)
    print(f"[{ts}] {event_type:10} | {name} ({ext}) | {size_kb} KB | {drive}")


def register_folder(wb, filepath, folder_path):
    """Log a newly discovered brand assets folder."""
    drive = os.path.splitdrive(folder_path)[0] or "?"
    ts    = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        count = sum(1 for _ in Path(folder_path).rglob("*") if _.is_file())
    except Exception:
        count = 0

    with _lock:
        ws = wb["Watched Folders"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == folder_path:
                return  # already registered
        n = ws.max_row
        ws.append([n, drive, folder_path, ts, count])
        wb.save(filepath)
    print(f"  [Watch] {folder_path} ({count} existing files)")


# ─── Watchdog Handler ─────────────────────────────────────────────────────────

class BrandAssetsHandler(FileSystemEventHandler):
    def __init__(self, wb, filepath, watched_folder):
        self.wb             = wb
        self.filepath       = filepath
        self.watched_folder = watched_folder

    def on_created(self, event):
        if not event.is_directory:
            save_event(self.wb, self.filepath, "ADDED",    event.src_path,  self.watched_folder)

    def on_moved(self, event):
        if not event.is_directory:
            save_event(self.wb, self.filepath, "MOVED IN", event.dest_path, self.watched_folder)

    def on_deleted(self, event):
        if not event.is_directory:
            save_event(self.wb, self.filepath, "DELETED",  event.src_path,  self.watched_folder)

    def on_modified(self, event):
        if not event.is_directory:
            save_event(self.wb, self.filepath, "MODIFIED", event.src_path,  self.watched_folder)


# ─── Drive Scanner ────────────────────────────────────────────────────────────

def get_available_drives():
    drives = []
    for letter in string.ascii_uppercase:
        drive = f"{letter}:\\"
        if os.path.exists(drive):
            drives.append(drive)
    return drives


def find_brand_asset_folders(drives):
    found     = []
    skip_dirs = {
        "windows", "system32", "syswow64", "program files",
        "$recycle.bin", "node_modules", ".git", "__pycache__",
        "appdata\\local\\temp"
    }
    for drive in drives:
        print(f"  Scanning {drive} ...")
        try:
            for root, dirs, _ in os.walk(drive, topdown=True):
                dirs[:] = [
                    d for d in dirs
                    if not any(s in os.path.join(root, d).lower() for s in skip_dirs)
                ]
                for d in dirs:
                    if d.lower() == FOLDER_NAME:
                        found.append(os.path.join(root, d))
        except PermissionError:
            pass
        except Exception as e:
            print(f"  Warning: {e}")
    return found


# ─── Watcher Manager ─────────────────────────────────────────────────────────

class WatcherManager:
    def __init__(self, wb, filepath):
        self.wb        = wb
        self.filepath  = filepath
        self.observers = {}   # folder_path -> Observer

    def add_folder(self, folder_path):
        if folder_path in self.observers:
            return
        register_folder(self.wb, self.filepath, folder_path)
        handler  = BrandAssetsHandler(self.wb, self.filepath, folder_path)
        observer = Observer()
        observer.schedule(handler, folder_path, recursive=True)
        observer.start()
        self.observers[folder_path] = observer

    def stop_all(self):
        for obs in self.observers.values():
            obs.stop(); obs.join()


def rescan_loop(manager, interval_minutes):
    while True:
        time.sleep(interval_minutes * 60)
        print(f"\n[{datetime.now().strftime('%H:%M:%S')}] Rescanning drives for new 'brand assets' folders...")
        for f in find_brand_asset_folders(get_available_drives()):
            manager.add_folder(f)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Brand Assets Folder Watcher")
    parser.add_argument("--output", default=DEFAULT_XLS,    help="Excel log path")
    parser.add_argument("--rescan", default=RESCAN_MINUTES, type=int,
                        help="Drive rescan interval in minutes (default: 30)")
    args = parser.parse_args()

    print("\n Brand Assets Watcher")
    print("=" * 50)
    print(f"  Target folder : '{FOLDER_NAME}' (any drive, any depth)")
    print(f"  Log file      : {args.output}")
    print(f"  Drive rescan  : every {args.rescan} minutes")
    print()

    wb      = init_excel(args.output)
    manager = WatcherManager(wb, args.output)

    print("Scanning all drives...")
    drives  = get_available_drives()
    print(f"  Drives: {', '.join(drives)}")
    folders = find_brand_asset_folders(drives)

    if folders:
        print(f"\nFound {len(folders)} 'brand assets' folder(s):")
        for f in folders:
            manager.add_folder(f)
    else:
        print("  No 'brand assets' folders found yet. Will rescan every "
              f"{args.rescan} minutes.")

    threading.Thread(target=rescan_loop, args=(manager, args.rescan), daemon=True).start()

    print(f"\nWatching {len(manager.observers)} folder(s). "
          "Press Ctrl+C to stop.\n")

    try:
        while True:
            time.sleep(2)
    except KeyboardInterrupt:
        print("\nStopped.")
        manager.stop_all()


if __name__ == "__main__":
    main()
