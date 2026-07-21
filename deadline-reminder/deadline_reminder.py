"""
deadline_reminder.py
--------------------
Daily deadline reminder system.
- Manages deadlines in D:\deadlines.xlsx (add/edit there)
- Sends Windows toast notifications at configured times
- Logs all reminders sent to D:\deadline_log.xlsx
- Auto-syncs deadlines.xlsx → deadlines.json every minute

Usage:
    python deadline_reminder.py                  # run reminder daemon
    python deadline_reminder.py --add            # interactive: add a deadline
    python deadline_reminder.py --list           # show all upcoming deadlines
    python deadline_reminder.py --test           # fire a test notification now
"""

import os
import sys
import json
import time
import argparse
import threading
from datetime import datetime, date, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl import load_workbook

# Notification times (24hr) — reminders fire at these times each day
REMINDER_TIMES = ["08:00", "13:00", "18:00"]

DEADLINES_XLS  = r"D:\deadlines.xlsx"
DEADLINES_JSON = r"D:\deadlines.json"
LOG_XLS        = r"D:\deadline_log.xlsx"

# ─── Urgency colors ──────────────────────────────────────────────────────────
COLORS = {
    "overdue":  "FF0000",  # red
    "today":    "FF6600",  # orange
    "tomorrow": "FFC000",  # amber
    "3days":    "FFFF00",  # yellow
    "week":     "92D050",  # green
    "future":   "DEEAF1",  # blue-grey
}

def urgency_color(deadline_date):
    today = date.today()
    delta = (deadline_date - today).days
    if delta < 0:    return COLORS["overdue"]
    if delta == 0:   return COLORS["today"]
    if delta == 1:   return COLORS["tomorrow"]
    if delta <= 3:   return COLORS["3days"]
    if delta <= 7:   return COLORS["week"]
    return COLORS["future"]

def urgency_label(deadline_date):
    today = date.today()
    delta = (deadline_date - today).days
    if delta < 0:    return f"OVERDUE by {abs(delta)} day(s)!"
    if delta == 0:   return "DUE TODAY"
    if delta == 1:   return "Due TOMORROW"
    if delta <= 7:   return f"Due in {delta} days"
    return f"Due in {delta} days"

# ─── Excel: Deadlines ────────────────────────────────────────────────────────

def init_deadlines_excel(filepath):
    """Create deadlines Excel if not present, with sample entries."""
    if not os.path.exists(filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Deadlines"
        headers  = ["Title", "Deadline Date", "Category", "Priority",
                    "Notes", "Remind At (HH:MM)", "Done (YES/NO)"]
        ws.append(headers)

        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill; cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 40
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 14

        # Sample rows
        today    = date.today()
        tomorrow = today + timedelta(days=1)
        next_wk  = today + timedelta(days=7)
        samples  = [
            ["GST Filing", today.strftime("%Y-%m-%d"),     "Tax",      "HIGH",   "Monthly GST return",             "08:00,13:00,18:00", "NO"],
            ["Brand Assets Update", tomorrow.strftime("%Y-%m-%d"), "Design", "MEDIUM", "Upload new logo variants", "08:00",             "NO"],
            ["Follow up with client", next_wk.strftime("%Y-%m-%d"), "Sales", "LOW",   "Vipin Global Advertising",  "09:00",             "NO"],
        ]
        for s in samples:
            ws.append(s)

        wb.save(filepath)
        print(f"Created deadlines file: {filepath}")
        print("  -> Open it to add/edit your deadlines!")
    return load_workbook(filepath)


def load_deadlines(filepath):
    """Read all non-done deadlines from Excel, return list of dicts."""
    if not os.path.exists(filepath):
        init_deadlines_excel(filepath)

    wb       = load_workbook(filepath)
    ws       = wb.active
    deadlines = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        title, dl_date, category, priority, notes, remind_at, done = row

        if str(done).strip().upper() == "YES":
            continue

        # Parse date
        if isinstance(dl_date, datetime):
            dl_date = dl_date.date()
        elif isinstance(dl_date, str):
            try:
                dl_date = datetime.strptime(dl_date.strip(), "%Y-%m-%d").date()
            except Exception:
                continue
        elif not isinstance(dl_date, date):
            continue

        times = [t.strip() for t in str(remind_at or "08:00").split(",")]

        deadlines.append({
            "title":    str(title).strip(),
            "date":     dl_date,
            "category": str(category or "General").strip(),
            "priority": str(priority or "MEDIUM").strip(),
            "notes":    str(notes or "").strip(),
            "times":    times,
        })

    return deadlines


def sync_to_json(deadlines, json_path):
    """Write deadlines to JSON for quick access by other scripts."""
    data = [{**d, "date": d["date"].isoformat()} for d in deadlines]
    with open(json_path, "w") as f:
        json.dump(data, f, indent=2)


# ─── Notification ─────────────────────────────────────────────────────────────

def send_notification(title, message, urgency="normal"):
    """Send a Windows toast notification."""
    try:
        from winotify import Notification, audio
        icon = ""  # optional: path to an .ico file

        toast = Notification(
            app_id  = "Deadline Reminder",
            title   = title,
            msg     = message,
            duration= "long"
        )
        if urgency == "high":
            toast.set_audio(audio.Default, loop=False)
        toast.show()
    except Exception:
        # Fallback to plyer
        try:
            from plyer import notification
            notification.notify(
                title   = title,
                message = message,
                timeout = 10
            )
        except Exception as e:
            print(f"  [Notification fallback failed: {e}]")


# ─── Log ─────────────────────────────────────────────────────────────────────

_log_lock = threading.Lock()

def init_log(filepath):
    if not os.path.exists(filepath):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reminder Log"
        ws.append(["#", "Sent At", "Title", "Deadline", "Days Left", "Category", "Priority"])
        fill = PatternFill("solid", fgColor="1F4E79")
        font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = fill; cell.font = font
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 10
        wb.save(filepath)
    return load_workbook(filepath)


def log_reminder(wb, filepath, deadline):
    today     = date.today()
    days_left = (deadline["date"] - today).days
    ts        = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with _log_lock:
        ws  = wb["Reminder Log"]
        row = ws.max_row + 1
        ws.append([
            row - 1, ts,
            deadline["title"],
            deadline["date"].isoformat(),
            days_left,
            deadline["category"],
            deadline["priority"]
        ])
        color = urgency_color(deadline["date"])
        fill  = PatternFill("solid", fgColor=color)
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = fill
        wb.save(filepath)


# ─── Scheduler ───────────────────────────────────────────────────────────────

def should_remind_today(deadline):
    """Only remind if deadline is within 30 days or overdue."""
    delta = (deadline["date"] - date.today()).days
    return delta <= 30


def reminder_loop(log_wb, log_path):
    """
    Main loop: checks every minute if it's time to send reminders.
    Reloads deadlines from Excel each check so edits are picked up live.
    """
    fired_today = set()   # (title, time_str) — avoid duplicate fires per day

    print("Reminder daemon started.")
    print(f"  Deadlines file : {DEADLINES_XLS}")
    print(f"  Log file       : {log_path}")
    print(f"  Default times  : {', '.join(REMINDER_TIMES)}")
    print()

    last_day = date.today()

    while True:
        now      = datetime.now()
        today    = date.today()
        hhmm     = now.strftime("%H:%M")

        # Reset daily fire tracker at midnight
        if today != last_day:
            fired_today.clear()
            last_day = today

        deadlines = load_deadlines(DEADLINES_XLS)
        sync_to_json(deadlines, DEADLINES_JSON)

        for dl in deadlines:
            if not should_remind_today(dl):
                continue

            for t in dl["times"]:
                key = (dl["title"], t)
                if hhmm == t and key not in fired_today:
                    fired_today.add(key)
                    label   = urgency_label(dl["date"])
                    urgency = "high" if (dl["date"] - today).days <= 1 else "normal"

                    msg = f"{label}\n{dl['date'].strftime('%d %b %Y')}"
                    if dl["notes"]:
                        msg += f"\n{dl['notes'][:80]}"

                    print(f"[{hhmm}] REMINDER: {dl['title']} — {label}")
                    send_notification(
                        title   = f"{'⚠️ ' if urgency == 'high' else ''}{dl['title']}",
                        message = msg,
                        urgency = urgency
                    )
                    log_reminder(log_wb, log_path, dl)

        time.sleep(30)  # check every 30 seconds


# ─── CLI helpers ─────────────────────────────────────────────────────────────

def cmd_list():
    deadlines = load_deadlines(DEADLINES_XLS)
    today     = date.today()
    print(f"\n{'='*60}")
    print(f"  UPCOMING DEADLINES  ({today.strftime('%d %b %Y')})")
    print(f"{'='*60}")
    if not deadlines:
        print("  No active deadlines found.")
    for dl in sorted(deadlines, key=lambda d: d["date"]):
        label = urgency_label(dl["date"])
        print(f"  [{dl['priority']:6}] {dl['title'][:35]:<35} {dl['date'].strftime('%d %b %Y')}  {label}")
    print()


def cmd_test():
    send_notification(
        "Deadline Reminder — TEST",
        "If you see this, notifications are working!\nYour deadlines will appear like this.",
        urgency="normal"
    )
    print("Test notification sent!")


def cmd_add():
    print("\nAdd a new deadline:")
    title    = input("  Title         : ").strip()
    dl_date  = input("  Date (YYYY-MM-DD): ").strip()
    category = input("  Category      : ").strip() or "General"
    priority = input("  Priority (HIGH/MEDIUM/LOW): ").strip().upper() or "MEDIUM"
    notes    = input("  Notes         : ").strip()
    times    = input("  Remind at (e.g. 08:00,13:00): ").strip() or "08:00,13:00,18:00"

    wb = init_deadlines_excel(DEADLINES_XLS)
    ws = wb["Deadlines"] if "Deadlines" in wb.sheetnames else wb.active
    ws.append([title, dl_date, category, priority, notes, times, "NO"])

    color = urgency_color(datetime.strptime(dl_date, "%Y-%m-%d").date())
    row   = ws.max_row
    for col in range(1, 8):
        ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=color)

    wb.save(DEADLINES_XLS)
    print(f"\nAdded: '{title}' due {dl_date}\n")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Daily Deadline Reminder")
    parser.add_argument("--list",  action="store_true", help="List all deadlines")
    parser.add_argument("--add",   action="store_true", help="Add a new deadline")
    parser.add_argument("--test",  action="store_true", help="Send test notification")
    parser.add_argument("--log",   default=LOG_XLS,     help="Log Excel path")
    args = parser.parse_args()

    init_deadlines_excel(DEADLINES_XLS)

    if args.list:
        cmd_list(); return
    if args.add:
        cmd_add();  return
    if args.test:
        cmd_test(); return

    # Run daemon
    log_wb = init_log(args.log)
    reminder_loop(log_wb, args.log)


if __name__ == "__main__":
    main()
